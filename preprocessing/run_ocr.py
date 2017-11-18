import sys
import io
import os
import re
import json
from wand.image import Image
from PIL import Image as PImage
import pyocr
import pyocr.builders
from openpyxl import load_workbook
from multiprocessing import Pool
import datetime


tools = pyocr.get_available_tools()
if len(tools) == 0:
    print("No OCR tool found")
    sys.exit(1)
# The tools are returned in the recommended order of usage
tool = tools[0]
print("Will use tool '%s'" % (tool.get_name()))

langs = tool.get_available_languages()
if 'eng' not in langs:
    print("No eng")
    sys.exit(1)
lang = 'eng_best'

def ocr(filename):
    name, ext = os.path.splitext(filename)
    if ext.lower() == '.pdf':
        image_pdf = Image(filename=filename, resolution=300)
        image_jpeg = image_pdf.convert('jpeg')
        img = image_jpeg.make_blob()
        img = PImage.open(io.BytesIO(img))
    if ext.lower() in ['.jpg', '.jpeg', '.png']:
        img = PImage.open(filename)
    word_boxes = tool.image_to_string(
        img,
        lang=lang,
        builder=pyocr.builders.LineBoxBuilder(tesseract_layout=1)
    )
    X, Y = img.size
    res = []
    for line in word_boxes:
        if line.position[1][0] - line.position[0][0] < 0.01 * X:
            continue
        if line.position[1][1] - line.position[0][1] < 0.003 * X:
            continue
        if line.content.strip() == '':
            continue
        res.append(line)
    return res, X, Y

def get_index_from_name(filename):
    name, ext = os.path.splitext(filename)
    if ext.lower() not in ['.pdf', '.jpg', '.jpeg', '.png']:
        return None
    digits = re.sub("\D", "", name)
    if digits not in name:
        print "ERROR", filename
    if digits == '':
        return None
    return int(digits)

def read_xlsx(filename):
    wb = load_workbook(filename)
    ws = wb.active
    objs = {}
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == None:
            continue
        idx = int(re.sub('\D', '', ws.cell(row=1, column=col).value))
        obj = {}
        for row in range(2, ws.max_row + 1):
            key = ws.cell(row=row, column=2).value
            val = ws.cell(row=row, column=col).value
            if val != None:
                if type(val) is datetime.datetime:
                    val = str(val)
                obj[key] = val
        objs[idx] = obj
    return objs

def gen_output(filename):
    fax_id = get_index_from_name(filename)
    if fax_id == None or fax_id not in objs:
        return None
    res = {
        'set_id': set_id,
        'fax_id': fax_id,
        'attrs': objs[fax_id],
        'words': []
    }
    word_boxes, width, height = ocr(path + filename)
    res['img_width'] = width
    res['img_height'] = height
    for line in word_boxes:
        obj_line = {'p':line.position, 'w':[]}
        for box in line.word_boxes:
            obj_line['w'].append((box.content, box.position))
        res['words'].append(obj_line)
    print fax_id,
    sys.stdout.flush() 
    return res


base_path ='../image/'
for direc in os.listdir(base_path):
    path = base_path + direc + '/'
    if not os.path.isdir(path):
        continue
    set_id = int(re.sub('\D', '', direc))
    # if set_id != 1:
    #     continue
    xlsx = [s for s in os.listdir(path) if s.endswith('.xlsx') and 'DT' not in s]
    assert len(xlsx) == 1
    xlsx_filename = xlsx[0]
    objs = read_xlsx(path + xlsx_filename)
    print direc, len(objs)
    output = []
    pool = Pool(processes=4)
    output = pool.map(gen_output, os.listdir(path))
    output = [x for x in output if x is not None]
    output = sorted(output, key=lambda x:x['fax_id'])
    output_f = file('../data/%d.json' % set_id, 'w')
    for obj in output:
        output_f.write(json.dumps(obj)+'\n')
    output_f.close()
    print


